#!/usr/bin/env python3
"""
生成角色權限表 Excel 檔案
根據 Azure Portal 後端的 permissions.py 定義與各 API 路由的實際權限檢查
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ============================================================
# 資料定義
# ============================================================

ROLES = ["root", "admin", "user"]
ROLE_LABELS = {
    "root": "最高管理者 (root)",
    "admin": "一般管理者 (admin)",
    "user": "一般使用者 (user)",
}
ROLE_HIERARCHY = {"root": 3, "admin": 2, "user": 1}

ROLE_PERMISSIONS = {
    "root": [
        "view_announcements", "use_agents", "view_library", "chat_history",
        "manage_users", "manage_library", "manage_announcements",
        "manage_agent_permissions", "access_all_agents", "access_all_docs",
        "cross_country_logs",
    ],
    "admin": [
        "view_announcements", "use_agents", "view_library", "chat_history",
        "manage_users", "manage_library", "manage_announcements",
        "manage_agent_permissions", "access_all_agents", "access_all_docs",
    ],
    "user": [
        "view_announcements", "use_agents", "view_library", "chat_history",
    ],
}

# 權限說明
PERMISSION_DESCRIPTIONS = {
    "view_announcements": "檢視公告",
    "use_agents": "使用 AI Agent 對話",
    "view_library": "檢視圖書館文件",
    "chat_history": "檢視對話歷史紀錄",
    "manage_users": "使用者管理（新增/編輯/停用/角色指派）",
    "manage_library": "圖書館管理（新增/編輯/刪除館藏與文件、權限設定）",
    "manage_announcements": "公告管理（新增/編輯/刪除/發佈公告、上傳附件）",
    "manage_agent_permissions": "Agent 權限管理（發佈/下架 Agent、設定 ACL）",
    "access_all_agents": "存取所有 Agent（不受 ACL 限制）",
    "access_all_docs": "存取所有文件（不受文件權限限制）",
    "cross_country_logs": "跨國稽核日誌查看",
}

# 功能模組 → 對應的 API 操作 → 所需權限
FEATURE_MATRIX = [
    # (功能模組, 操作說明, 所需權限 key, 備註)
    ("首頁", "檢視首頁公告", "view_announcements", ""),
    ("AI Agent 商店", "瀏覽 Agent 列表", "use_agents", "依 ACL 過濾可見 Agent"),
    ("AI Agent 商店", "使用 Agent 對話（含串流）", "use_agents", ""),
    ("AI Agent 商店", "檢視對話歷史", "chat_history", ""),
    ("圖書館", "瀏覽圖書館館藏與文件", "view_library", "依文件 auth_rules 過濾"),
    ("設定 — 公告管理", "新增 / 編輯 / 刪除公告", "manage_announcements", "root 可跨國操作"),
    ("設定 — 公告管理", "發佈 / 取消發佈公告", "manage_announcements", ""),
    ("設定 — 公告管理", "上傳 / 刪除公告附件", "manage_announcements", ""),
    ("設定 — Agent 權限", "檢視所有 Agent（含未發佈）", "manage_agent_permissions", ""),
    ("設定 — Agent 權限", "發佈 / 下架 Agent", "manage_agent_permissions", ""),
    ("設定 — Agent 權限", "設定 Agent ACL（角色/使用者/例外）", "manage_agent_permissions", ""),
    ("設定 — 圖書館管理", "新增 / 編輯 / 刪除館藏目錄", "manage_library", "root 可跨國操作"),
    ("設定 — 圖書館管理", "新增 / 編輯 / 刪除文件", "manage_library", ""),
    ("設定 — 圖書館管理", "上傳 / 刪除文件附件", "manage_library", ""),
    ("設定 — 圖書館管理", "設定文件存取權限（角色/使用者/例外）", "manage_library", ""),
    ("設定 — 圖書館管理", "檢視圖書館統計", "manage_library", ""),
    ("設定 — 使用者管理", "檢視使用者列表", "manage_users", "root / admin 皆可看全部國家（不指定國家時不篩選）"),
    ("設定 — 使用者管理", "國家篩選下拉選單", "manage_users", "root / admin 皆可見"),
    ("設定 — 使用者管理", "新增使用者（可指定任意國家）", "manage_users", "root / admin 皆可選擇任意國家建立使用者"),
    ("設定 — 使用者管理", "編輯使用者資料", "manage_users", "不可操作 ≥ 自身等級的使用者；admin 跨國使用者為唯讀"),
    ("設定 — 使用者管理", "編輯使用者 — 修改國家欄位", "manage_users", "root / admin 皆可修改使用者國家（限本國使用者）"),
    ("設定 — 使用者管理", "變更使用者角色", "manage_users", "不可指派 ≥ 自身等級的角色；admin 跨國使用者為唯讀"),
    ("設定 — 使用者管理", "停用 / 啟用使用者", "manage_users", "不可操作 ≥ 自身等級的使用者；admin 跨國使用者為唯讀"),
    ("設定 — 使用者管理", "刪除使用者", "manage_users", "不可刪除 ≥ 自身等級的使用者；admin 跨國使用者為唯讀"),
    ("設定 — 稽核日誌", "檢視跨國稽核日誌", "cross_country_logs", "僅 root 擁有此權限"),
    ("設定 — 稽核日誌", "匯出稽核日誌", "cross_country_logs", ""),
]

# ============================================================
# 樣式定義
# ============================================================

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)

SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Microsoft JhengHei", bold=True, size=10)

YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YES_FONT = Font(name="Microsoft JhengHei", color="006100", bold=True, size=10)

NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NO_FONT = Font(name="Microsoft JhengHei", color="9C0006", size=10)

NORMAL_FONT = Font(name="Microsoft JhengHei", size=10)
BOLD_FONT = Font(name="Microsoft JhengHei", bold=True, size=10)

TITLE_FONT = Font(name="Microsoft JhengHei", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="Microsoft JhengHei", size=10, color="666666")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def apply_cell(cell, font=NORMAL_FONT, alignment=LEFT_WRAP):
    cell.font = font
    cell.alignment = alignment
    cell.border = THIN_BORDER


# ============================================================
# Sheet 1: 權限代碼總覽
# ============================================================

def create_permission_overview(wb: Workbook):
    ws = wb.active
    ws.title = "權限代碼總覽"

    # 標題
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Azure Portal — 角色權限總覽"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:F2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = "三角色設計：root（最高管理者）> admin（一般管理者）> user（一般使用者）"
    subtitle_cell.font = SUBTITLE_FONT

    # 表頭 (row 4)
    headers = ["權限代碼", "權限說明", "root\n最高管理者", "admin\n一般管理者", "user\n一般使用者", "備註"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header(cell)

    # 所有權限列表（按分類排序）
    all_permissions = list(PERMISSION_DESCRIPTIONS.keys())

    for row_idx, perm in enumerate(all_permissions, 5):
        # 權限代碼
        cell = ws.cell(row=row_idx, column=1, value=perm)
        apply_cell(cell, font=Font(name="Consolas", size=10))

        # 權限說明
        cell = ws.cell(row=row_idx, column=2, value=PERMISSION_DESCRIPTIONS[perm])
        apply_cell(cell)

        # 各角色是否擁有
        for role_col, role in enumerate(ROLES, 3):
            has = perm in ROLE_PERMISSIONS[role]
            cell = ws.cell(row=row_idx, column=role_col, value="✓" if has else "✗")
            if has:
                cell.fill = YES_FILL
                cell.font = YES_FONT
            else:
                cell.fill = NO_FILL
                cell.font = NO_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        # 備註
        note = ""
        if perm == "cross_country_logs":
            note = "僅 root 獨有"
        elif perm in ("access_all_agents", "access_all_docs"):
            note = "root / admin 共有，user 受 ACL 限制"
        elif perm == "manage_users":
            note = "root / admin 共有（皆可跨國查看；admin 跨國使用者為唯讀）"
        elif perm.startswith("manage_"):
            note = "root / admin 共有（root 可跨國操作）"
        cell = ws.cell(row=row_idx, column=6, value=note)
        apply_cell(cell)

    # 欄寬
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 40

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[4].height = 35


# ============================================================
# Sheet 2: 功能模組權限矩陣
# ============================================================

def create_feature_matrix(wb: Workbook):
    ws = wb.create_sheet("功能模組權限矩陣")

    # 標題
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Azure Portal — 功能模組權限矩陣"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:G2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = "每個功能操作對應的角色存取權限（✓ = 可存取, ✗ = 不可存取）"
    subtitle_cell.font = SUBTITLE_FONT

    # 表頭 (row 4)
    headers = ["功能模組", "操作說明", "所需權限", "root\n最高管理者", "admin\n一般管理者", "user\n一般使用者", "備註"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header(cell)

    current_module = ""
    row = 5
    for module, operation, perm_key, note in FEATURE_MATRIX:
        # 模組名稱（合併同模組的儲存格稍後處理，這裡先寫入）
        if module != current_module:
            cell = ws.cell(row=row, column=1, value=module)
            apply_cell(cell, font=BOLD_FONT)
            current_module = module
        else:
            cell = ws.cell(row=row, column=1, value="")
            apply_cell(cell)

        # 操作說明
        cell = ws.cell(row=row, column=2, value=operation)
        apply_cell(cell)

        # 所需權限
        cell = ws.cell(row=row, column=3, value=perm_key)
        apply_cell(cell, font=Font(name="Consolas", size=9))

        # 各角色
        for role_col, role in enumerate(ROLES, 4):
            has = perm_key in ROLE_PERMISSIONS[role]
            cell = ws.cell(row=row, column=role_col, value="✓" if has else "✗")
            if has:
                cell.fill = YES_FILL
                cell.font = YES_FONT
            else:
                cell.fill = NO_FILL
                cell.font = NO_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        # 備註
        cell = ws.cell(row=row, column=7, value=note)
        apply_cell(cell)

        row += 1

    # 欄寬
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 35

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[4].height = 35


# ============================================================
# Sheet 3: 角色階層與操作規則
# ============================================================

def create_role_hierarchy(wb: Workbook):
    ws = wb.create_sheet("角色階層與操作規則")

    # 標題
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Azure Portal — 角色階層與操作規則"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- 角色階層表 ---
    ws.merge_cells("A3:D3")
    cell = ws["A3"]
    cell.value = "一、角色階層"
    cell.font = Font(name="Microsoft JhengHei", bold=True, size=12, color="1F4E79")

    headers = ["角色代碼", "角色名稱", "階層等級", "說明"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header(cell)

    role_data = [
        ("root", "最高管理者", 3, "系統最高權限，可跨國操作、查看稽核日誌"),
        ("admin", "一般管理者", 2, "管理使用者、Agent、圖書館、公告；新增/編輯使用者時可指定任意國家"),
        ("user", "一般使用者", 1, "僅可使用基本功能（瀏覽、對話、查看歷史）"),
    ]
    for row_idx, (code, name, level, desc) in enumerate(role_data, 5):
        ws.cell(row=row_idx, column=1, value=code).font = Font(name="Consolas", size=10)
        ws.cell(row=row_idx, column=2, value=name).font = NORMAL_FONT
        ws.cell(row=row_idx, column=3, value=level).font = NORMAL_FONT
        ws.cell(row=row_idx, column=3).alignment = CENTER
        ws.cell(row=row_idx, column=4, value=desc).font = NORMAL_FONT
        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
            ws.cell(row=row_idx, column=c).alignment = LEFT_WRAP if c != 3 else CENTER

    # --- 操作規則 ---
    ws.merge_cells("A9:D9")
    cell = ws["A9"]
    cell.value = "二、角色操作規則（使用者管理）"
    cell.font = Font(name="Microsoft JhengHei", bold=True, size=12, color="1F4E79")

    rules = [
        ("規則 1", "不可操作自己", "任何管理者都不能修改自己的帳號設定，需聯繫上級管理者"),
        ("規則 2", "不可操作等級 ≥ 自己的使用者", "例如 admin 不能編輯/停用 admin 或 root"),
        ("規則 3", "不可指派等級 ≥ 自己的角色", "例如 admin 只能指派 user 角色，不能指派 admin 或 root"),
    ]

    headers2 = ["規則", "摘要", "詳細說明"]
    for col_idx, header in enumerate(headers2, 1):
        cell = ws.cell(row=10, column=col_idx, value=header)
        apply_header(cell)

    for row_idx, (rule, summary, detail) in enumerate(rules, 11):
        ws.cell(row=row_idx, column=1, value=rule).font = BOLD_FONT
        ws.cell(row=row_idx, column=2, value=summary).font = NORMAL_FONT
        ws.cell(row=row_idx, column=3, value=detail).font = NORMAL_FONT
        for c in range(1, 4):
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
            ws.cell(row=row_idx, column=c).alignment = LEFT_WRAP

    # --- 可指派角色矩陣 ---
    ws.merge_cells("A15:D15")
    cell = ws["A15"]
    cell.value = "三、可指派角色矩陣"
    cell.font = Font(name="Microsoft JhengHei", bold=True, size=12, color="1F4E79")

    headers3 = ["操作者角色", "可指派的角色", "可操作的目標使用者", "說明"]
    for col_idx, header in enumerate(headers3, 1):
        cell = ws.cell(row=16, column=col_idx, value=header)
        apply_header(cell)

    assign_data = [
        ("root", "admin, user", "admin, user", "可管理所有非 root 使用者"),
        ("admin", "user", "user", "僅可管理 user 角色的使用者"),
        ("user", "（無）", "（無）", "無管理權限"),
    ]
    for row_idx, (role, assignable, operable, desc) in enumerate(assign_data, 17):
        ws.cell(row=row_idx, column=1, value=role).font = Font(name="Consolas", bold=True, size=10)
        ws.cell(row=row_idx, column=2, value=assignable).font = NORMAL_FONT
        ws.cell(row=row_idx, column=3, value=operable).font = NORMAL_FONT
        ws.cell(row=row_idx, column=4, value=desc).font = NORMAL_FONT
        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
            ws.cell(row=row_idx, column=c).alignment = LEFT_WRAP

    # --- root vs admin 差異 ---
    ws.merge_cells("A21:D21")
    cell = ws["A21"]
    cell.value = "四、root 與 admin 的關鍵差異"
    cell.font = Font(name="Microsoft JhengHei", bold=True, size=12, color="1F4E79")

    headers4 = ["功能", "root", "admin", "差異說明"]
    for col_idx, header in enumerate(headers4, 1):
        cell = ws.cell(row=22, column=col_idx, value=header)
        apply_header(cell)

    diff_data = [
        ("查看使用者列表（跨國）", "✓ 可看全部國家", "✓ 可看全部國家", "root / admin 不指定國家時皆可看全部"),
        ("新增使用者 — 選擇國家", "✓ 可指定任意國家", "✓ 可指定任意國家", "root / admin 皆可在新增使用者時選擇任意國家"),
        ("編輯使用者 — 修改國家", "✓ 可修改", "✓ 可修改（限本國使用者）", "admin 可修改本國使用者的國家；跨國使用者為唯讀"),
        ("編輯/停用/刪除跨國使用者", "✓ 可操作", "✗ 唯讀（可看不可改）", "admin 可看到跨國使用者但操作按鈕被禁用"),
        ("國家篩選下拉選單", "✓ 顯示", "✓ 顯示", "前端 root / admin 皆可見國家篩選"),
        ("跨國操作（公告/圖書館）", "✓ 可指定任意國家", "✗ 僅限本國", "root 可透過 country 參數操作其他國家的公告與圖書館"),
        ("稽核日誌", "✓ 可查看", "✗ 不可查看", "cross_country_logs 權限僅 root 擁有"),
        ("管理 admin 使用者", "✓ 可管理", "✗ 不可管理", "admin 無法操作等級 ≥ 自己的使用者"),
        ("指派 admin 角色", "✓ 可指派", "✗ 不可指派", "admin 無法指派等級 ≥ 自己的角色"),
        ("側邊欄稽核日誌入口", "✓ 顯示", "✗ 隱藏", "前端依 cross_country_logs 權限控制顯示"),
    ]
    for row_idx, (feature, root_val, admin_val, desc) in enumerate(diff_data, 23):
        ws.cell(row=row_idx, column=1, value=feature).font = BOLD_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).alignment = LEFT_WRAP

        cell_root = ws.cell(row=row_idx, column=2, value=root_val)
        cell_root.border = THIN_BORDER
        cell_root.alignment = CENTER
        if root_val.startswith("✓"):
            cell_root.fill = YES_FILL
            cell_root.font = YES_FONT
        else:
            cell_root.fill = NO_FILL
            cell_root.font = NO_FONT

        cell_admin = ws.cell(row=row_idx, column=3, value=admin_val)
        cell_admin.border = THIN_BORDER
        cell_admin.alignment = CENTER
        if admin_val.startswith("✓"):
            cell_admin.fill = YES_FILL
            cell_admin.font = YES_FONT
        else:
            cell_admin.fill = NO_FILL
            cell_admin.font = NO_FONT

        ws.cell(row=row_idx, column=4, value=desc).font = NORMAL_FONT
        ws.cell(row=row_idx, column=4).border = THIN_BORDER
        ws.cell(row=row_idx, column=4).alignment = LEFT_WRAP

    # 欄寬
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 50

    ws.row_dimensions[1].height = 35


# ============================================================
# Sheet 4: 前端頁面存取控制
# ============================================================

def create_frontend_access(wb: Workbook):
    ws = wb.create_sheet("前端頁面存取控制")

    # 標題
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Azure Portal — 前端頁面存取控制"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:F2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = "側邊欄選單項目依使用者權限動態顯示/隱藏"
    subtitle_cell.font = SUBTITLE_FONT

    headers = ["頁面 / 選單項目", "路由路徑", "所需權限", "root", "admin", "user"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header(cell)

    pages = [
        ("首頁", "/", "（無，所有人可見）", True, True, True),
        ("Agent 商店", "/agent-store", "use_agents", True, True, True),
        ("Agent 對話", "/agent-store/chat", "use_agents", True, True, True),
        ("對話歷史", "/agent-store/history", "chat_history", True, True, True),
        ("圖書館", "/library", "view_library", True, True, True),
        ("設定 — 公告管理", "/settings/announcements", "manage_announcements", True, True, False),
        ("設定 — Agent 權限", "/settings/agent-permissions", "manage_agent_permissions", True, True, False),
        ("設定 — 圖書館管理", "/settings/library", "manage_library", True, True, False),
        ("設定 — 圖書館統計", "/settings/library-stats", "manage_library", True, True, False),
        ("設定 — 使用者管理", "/settings/users", "manage_users", True, True, False),
        ("設定 — 稽核日誌", "/settings/audit-logs", "cross_country_logs", True, False, False),
    ]

    for row_idx, (page, route, perm, root_ok, admin_ok, user_ok) in enumerate(pages, 5):
        ws.cell(row=row_idx, column=1, value=page).font = NORMAL_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).alignment = LEFT_WRAP

        ws.cell(row=row_idx, column=2, value=route).font = Font(name="Consolas", size=9)
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).alignment = LEFT_WRAP

        ws.cell(row=row_idx, column=3, value=perm).font = Font(name="Consolas", size=9)
        ws.cell(row=row_idx, column=3).border = THIN_BORDER
        ws.cell(row=row_idx, column=3).alignment = LEFT_WRAP

        for col_offset, has_access in enumerate([root_ok, admin_ok, user_ok], 4):
            cell = ws.cell(row=row_idx, column=col_offset, value="✓" if has_access else "✗")
            if has_access:
                cell.fill = YES_FILL
                cell.font = YES_FONT
            else:
                cell.fill = NO_FILL
                cell.font = NO_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # 欄寬
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[4].height = 30


# ============================================================
# 主程式
# ============================================================

def main():
    wb = Workbook()

    create_permission_overview(wb)
    create_feature_matrix(wb)
    create_role_hierarchy(wb)
    create_frontend_access(wb)

    output_path = "角色權限表.xlsx"
    wb.save(output_path)
    print(f"✅ 角色權限表已生成：{output_path}")


if __name__ == "__main__":
    main()
